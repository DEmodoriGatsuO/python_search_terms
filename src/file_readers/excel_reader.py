import logging
from openpyxl import load_workbook
from pyxlsb import open_workbook as open_xlsb
import xlrd

def read_excel(file_path):
    try:
        if file_path.endswith(('.xlsx', '.xlsm')):  # .xlsm を追加
            return read_xlsx_xlsm(file_path)
        elif file_path.endswith(('.xls', '.xlsb')):
            return read_xls_xlsb(file_path)
    except Exception as e:
        logging.error(f"Error reading Excel file {file_path}: {e}")
        return ""

def read_xlsx_xlsm(file_path):
    wb = load_workbook(file_path, read_only=True, data_only=True)  # data_only=True を追加
    all_sheets_text = []
    for sheet in wb:
        sheet_text = '\n'.join(
            ' '.join(str(cell.value) if cell.value is not None else '' for cell in row)
            for row in sheet.iter_rows()
        )
        all_sheets_text.append(f"Sheet {sheet.title}:\n{sheet_text}")
    return '\n\n'.join(all_sheets_text)

def read_xls_xlsb(file_path):
    if file_path.endswith('.xlsb'):
        with open_xlsb(file_path) as wb:
            all_sheets_text = []
            for sheet in wb.sheets:
                sheet_text = '\n'.join(' '.join(str(cell.v) if cell.v is not None else '' for cell in row) for row in sheet)
                all_sheets_text.append(f"Sheet {sheet.name}:\n{sheet_text}")
            return '\n\n'.join(all_sheets_text)
    else:  # .xls の場合
        wb = xlrd.open_workbook(file_path)
        all_sheets_text = []
        for sheet in wb.sheets():
            sheet_text = '\n'.join(' '.join(str(cell.value) if cell.value else '' for cell in row) for row in sheet.get_rows())
            all_sheets_text.append(f"Sheet {sheet.name}:\n{sheet_text}")
        return '\n\n'.join(all_sheets_text)